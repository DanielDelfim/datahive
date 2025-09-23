# scripts/reposicao/exportar_matriz_xlsx.py
from __future__ import annotations

import argparse
import json
import shutil
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List

# Fonte única de paths do módulo de reposição
from app.utils.reposicao.config import RESULTS_DIR  # deve apontar para C:\Apps\Datahive\data\marketplaces\meli\reposicao\results

# ----------------- helpers base -----------------

def _parse_list_csv(s: str | None, default: Iterable[str]) -> List[str]:
    if not s:
        return list(default)
    return [x.strip() for x in s.split(",") if x.strip()]

def _load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def _atomic_replace(tmp: Path, target: Path, retries: int = 5, wait_secs: float = 0.6) -> None:
    """
    Substitui target por tmp com tentativas (contorna locks do Windows).
    Se falhar, salva como <target>.NEW.
    """
    bak = target.with_suffix(target.suffix + ".bak")
    if target.exists():
        try:
            if bak.exists():
                bak.unlink()
            target.replace(bak)
        except Exception:
            pass

    last_err: Exception | None = None
    for _ in range(max(1, retries)):
        try:
            tmp.replace(target)
            return
        except (PermissionError, OSError) as e:
            last_err = e
            time.sleep(wait_secs)

    alt = target.with_suffix(target.suffix + ".NEW")
    try:
        if alt.exists():
            alt.unlink()
        shutil.move(str(tmp), str(alt))
        print(f"[WARN] Não foi possível substituir {target}. Salvo como: {alt}")
    except Exception as e:
        raise last_err or e

def _num(x: Any) -> float:
    try:
        return float(x)
    except Exception:
        return 0.0

def _title(rec: Dict[str, Any]) -> str:
    t = (
        rec.get("title")
        or rec.get("titulo")
        or rec.get("nome")
        or rec.get("name")
        or (rec.get("raw") or {}).get("title")
        or (rec.get("raw") or {}).get("name")
        or ""
    )
    return str(t).strip()

def _coerce_items(items: Any) -> Dict[str, Dict[str, Any]]:
    """
    Aceita dict {gtin: rec} OU lista [ {gtin: "...", ...}, ... ].
    Retorna dict indexado por GTIN (string).
    """
    if isinstance(items, dict):
        out: Dict[str, Dict[str, Any]] = {}
        for gtin, rec in items.items():
            if isinstance(rec, dict):
                out[str(gtin)] = rec
        return out

    if isinstance(items, list):
        out: Dict[str, Dict[str, Any]] = {}
        for rec in items:
            if not isinstance(rec, dict):
                continue
            gtin = str(rec.get("gtin") or "").strip()
            if gtin:
                out[gtin] = rec
        return out

    return {}

# ----------------- Excel -----------------

def _write_sheet(ws, registros: Dict[str, Any]) -> None:
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "GTIN", "Título",
        "Vendidos 7d", "Vendidos 15d", "Vendidos 30d",
        "Média 7d", "Média 15d", "Média 30d",
        "Média preferida", "Média ponderada",
        "Estoque (matriz)",
        "Expectativa 30d", "Expectativa 60d",
        "Reposição 30d", "Reposição 60d",
    ]
    ws.title = "MATRIZ"
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ordena por título (asc), caindo para gtin
    ordered = sorted(registros.items(), key=lambda kv: (_title(kv[1]).lower(), kv[0]))

    sem_titulo = 0
    exemplos = []

    for gtin, rec in ordered:
        t = _title(rec)
        if not t:
            sem_titulo += 1
            if len(exemplos) < 3:
                exemplos.append(gtin)

        row = [
            gtin,
            t,
            _num(rec.get("sold_7")),
            _num(rec.get("sold_15")),
            _num(rec.get("sold_30")),
            _num(rec.get("media_diaria_7")),
            _num(rec.get("media_diaria_15")),
            _num(rec.get("media_diaria_30")),
            _num(rec.get("media_diaria_preferida")),
            _num(rec.get("media_diaria_ponderada")),
            _num(rec.get("estoque_matriz") or rec.get("estoque")),
            _num(rec.get("expectativa_30d")),
            _num(rec.get("expectativa_60d")),
            _num(rec.get("reposicao_necessaria_30d") or rec.get("reposicao_30d")),
            _num(rec.get("reposicao_necessaria_60d") or rec.get("reposicao_60d")),
        ]
        ws.append(row)

    widths = [16, 60, 12, 12, 12, 12, 12, 12, 16, 16, 16, 16, 16, 16, 16]
    align_right_cols = set(range(3, len(headers)+1))
    for idx, w in enumerate(widths, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = w
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal=("right" if idx in align_right_cols else "left"),
                vertical="center"
            )

    if sem_titulo:
        print(f"[WARN] {sem_titulo} itens sem título. Exemplos GTIN: {', '.join(exemplos)}")

def _save_workbook_atomic(wb, target: Path, atomic: bool = True) -> None:
    _ensure_parent(target)
    tmp = target.with_suffix(target.suffix + ".tmp")
    wb.save(tmp)
    if atomic:
        _atomic_replace(tmp, target)
    else:
        if target.exists():
            try:
                target.unlink()
            except Exception:
                pass
        shutil.move(str(tmp), str(target))

# ----------------- CLI -----------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Exporta reposição da MATRIZ (por GTIN) para Excel a partir de reposicao_matriz.json"
    )
    ap.add_argument("--infile", help="JSON base. Default: RESULTS_DIR()/reposicao_matriz.json")
    ap.add_argument("--outfile", help="Default: RESULTS_DIR()/reposicao_matriz.xlsx")
    ap.add_argument("--no-atomic", action="store_true", help="Desativa escrita atômica (save direto)")

    args = ap.parse_args()

    base = RESULTS_DIR()  # esperado: C:\\Apps\\Datahive\\data\\marketplaces\\meli\\reposicao\\results
    infile = Path(args.infile) if args.infile else (base / "reposicao_matriz.json")
    outfile = Path(args.outfile) if args.outfile else (base / "reposicao_matriz.xlsx")

    data = _load_json(infile)
    items = _coerce_items(data.get("items"))

    if not items:
        print("[WARN] Nada a exportar: 'items' vazio ou estrutura inesperada.")
        return

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    _write_sheet(ws, items)
    _save_workbook_atomic(wb, outfile, atomic=not args.no_atomic)
    print(f"[OK] Matriz por GTIN → {outfile}")

if __name__ == "__main__":
    main()
