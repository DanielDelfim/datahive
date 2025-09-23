from __future__ import annotations

import argparse
import json
from pathlib import Path
import time
import shutil
from typing import Any, Dict, Iterable, List, Tuple

from app.utils.reposicao.config import RESULTS_DIR

# Dependência: openpyxl (usual em ambiente de dados). Se preferir pandas, dá pra adaptar.


# ----------------- helpers base -----------------

def _parse_list_csv(s: str | None, default: Iterable[str]) -> List[str]:
    if not s:
        return list(default)
    return [x.strip().lower() for x in s.split(",") if x.strip()]

def _load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def _ensure_parent(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def _atomic_replace(tmp: Path, target: Path, retries: int = 5, wait_secs: float = 0.6) -> None:
    """
    Tenta substituir target por tmp, com retries para contornar lock no Windows (Excel/antivírus).
    Se ainda falhar, salva como <target>.NEW e informa no log.
    """
    bak = target.with_suffix(target.suffix + ".bak")
    # tenta criar .bak (best-effort)
    if target.exists():
        try:
            if bak.exists():
                bak.unlink()
            target.replace(bak)
        except Exception:
            # ignora falha no .bak (não bloqueia a troca)
            pass

    last_err: Exception | None = None
    for _ in range(max(1, retries)):
        try:
            tmp.replace(target)  # os.replace
            return
        except PermissionError as e:
            last_err = e
            time.sleep(wait_secs)
        except OSError as e:
            # Alguns antivírus retornam códigos genéricos; faz retry
            last_err = e
            time.sleep(wait_secs)

    # Fallback: mantém arquivo alternativo para não perder saída
    alt = target.with_suffix(target.suffix + ".NEW")
    try:
        if alt.exists():
            alt.unlink()
        shutil.move(str(tmp), str(alt))
        print(f"[WARN] Não foi possível substituir {target} (arquivo pode estar aberto).")
        print(f"       Salvo como arquivo alternativo: {alt}")
        print("       Feche o arquivo de destino/Excel e renomeie o .NEW para o nome final.")
    except Exception as e:
        # último recurso: levanta erro original para não silenciar
        raise last_err or e

def _sort_key(kind: str, item: Tuple[str, Dict[str, Any]]):
    mlb, rec = item
    if kind == "mlb":
        return mlb
    if kind == "sold_30":
        try:
            # decrescente por sold_30
            return -float(rec.get("sold_30", 0) or 0.0)
        except Exception:
            return 0
    # default: title
    return str(rec.get("title") or rec.get("nome") or "").lower()


# ----------------- Excel helpers -----------------

def _write_sheet(ws, loja: str, registros: Dict[str, Any], sort_by: str) -> None:
    """
    Escreve uma aba com cabeçalho e linhas formatadas.
+    Colunas: MLB | Título | Sold 7 | Sold 15 | Sold 30 | Estoque | Rep 30d | Rep 60d
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "MLB", "Título", "Vendidos 7d", "Vendidos 15d", "Vendidos 30d",
        "Estoque", "Reposição 30d", "Reposição 60d"
    ]
    ws.title = loja.upper()

    # Cabeçalho
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linhas
    items = list(registros.items())
    items.sort(key=lambda it: _sort_key(sort_by, it))

    for mlb, rec in items:
        if not isinstance(rec, dict):
            continue
        row = [
            mlb,
            str(rec.get("title") or rec.get("nome") or ""),
            _num(rec.get("sold_7")),
            _num(rec.get("sold_15")),
            _num(rec.get("sold_30")),
            _num(rec.get("estoque")),
            _num(rec.get("reposicao_necessaria_30d")),
            _num(rec.get("reposicao_necessaria_60d")),
        ]
        ws.append(row)

    # Largura de colunas e alinhamento
    widths = [16, 60, 12, 12, 12, 12, 16, 16]
    align_right_cols = {3, 4, 5, 6, 7, 8}
    for idx, w in enumerate(widths, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = w
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            if idx in align_right_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

def _num(x: Any) -> float:
    try:
        return float(x)
    except Exception:
        return 0.0

def _save_workbook_atomic(wb, target: Path, atomic: bool = True) -> None:
    _ensure_parent(target)
    tmp = target.with_suffix(target.suffix + ".tmp")
    wb.save(tmp)
    if atomic:
        _atomic_replace(tmp, target)
    else:
        # Save direto (não atômico) — usar apenas se necessário
        if target.exists():
            try:
                target.unlink()
            except Exception:
                pass
        shutil.move(str(tmp), str(target))


# ----------------- principal -----------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Exporta resumos SP/MG em Excel a partir de vendas_7_15_30.json"
    )
    ap.add_argument("--infile", help="JSON base. Default: RESULTS_DIR()/vendas_7_15_30.json")
    ap.add_argument("--lojas", help="Ex.: SP,MG. Default: SP,MG")
    ap.add_argument("--outfile-sp", help="Default: RESULTS_DIR()/resumo_sp.xlsx")
    ap.add_argument("--outfile-mg", help="Default: RESULTS_DIR()/resumo_mg.xlsx")
    ap.add_argument("--sort", choices=["title", "mlb", "sold_30"], default="title",
                    help="Ordenação nas listas (default: title). 'sold_30' = desc.")
    ap.add_argument("--single-file", help="Se informado, cria um único arquivo com abas SP e MG.")
    ap.add_argument("--no-atomic", action="store_true", help="Desativa escrita atômica (usa save direto).")

    args = ap.parse_args()

    base_dir = RESULTS_DIR()
    infile = Path(args.infile) if args.infile else (base_dir / "vendas_7_15_30.json")
    outfile_sp = Path(args.outfile_sp) if args.outfile_sp else (base_dir / "resumo_sp.xlsx")
    outfile_mg = Path(args.outfile_mg) if args.outfile_mg else (base_dir / "resumo_mg.xlsx")

    lojas = _parse_list_csv(args.lojas, default=("sp", "mg")) if args.lojas else ["sp", "mg"]

    payload = _load_json(infile)
    results = payload.get("results") or {}
    if not isinstance(results, dict):
        print("[ERRO] Estrutura inválida em 'results'.")
        return

    if args.single_file:
        from openpyxl import Workbook
        wb = Workbook()
        # Remove aba padrão criada automaticamente
        default = wb.active
        wb.remove(default)

        wrote_any = False
        for loja in lojas:
            loja_key = loja.lower()
            registros = results.get(loja_key)
            if not isinstance(registros, dict):
                print(f"[WARN] Loja {loja_key.upper()} não encontrada em 'results'. Ignorando.")
                continue
            ws = wb.create_sheet(loja_key.upper())
            _write_sheet(ws, loja_key, registros, sort_by=args.sort)
            wrote_any = True

        if not wrote_any:
            print("[WARN] Nenhuma loja válida encontrada. Nada a exportar.")
            return

        target = Path(args.single_file)
        _save_workbook_atomic(wb, target, atomic=not args.no_atomic)
        print(f"[OK] Resumo consolidado → {target}")
        return

    # Dois arquivos (SP e MG)
    if "sp" in [loja_name.lower() for loja_name in lojas]:
        sp_data = results.get("sp")
        if isinstance(sp_data, dict):
            from openpyxl import Workbook
            wb_sp = Workbook()
            ws_sp = wb_sp.active
            _write_sheet(ws_sp, "sp", sp_data, sort_by=args.sort)
            _save_workbook_atomic(wb_sp, outfile_sp, atomic=not args.no_atomic)
            print(f"[OK] resumo SP → {outfile_sp}")
        else:
            print("[WARN] Loja SP não encontrada em 'results'.")

    if "mg" in [loja_name.lower() for loja_name in lojas]:
        mg_data = results.get("mg")
        if isinstance(mg_data, dict):
            from openpyxl import Workbook
            wb_mg = Workbook()
            ws_mg = wb_mg.active
            _write_sheet(ws_mg, "mg", mg_data, sort_by=args.sort)
            _save_workbook_atomic(wb_mg, outfile_mg, atomic=not args.no_atomic)
            print(f"[OK] resumo MG → {outfile_mg}")
        else:
            print("[WARN] Loja MG não encontrada em 'results'.")


if __name__ == "__main__":
    main()
