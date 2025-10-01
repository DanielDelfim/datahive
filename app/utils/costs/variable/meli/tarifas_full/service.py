from typing import Optional, List
import warnings
import pandas as pd
from openpyxl import load_workbook
from .mapper import map_aba1, map_aba2, map_aba3, map_aba4

from .aggregator import (
    enrich_and_clean_aba1, enrich_and_clean_aba2,
    enrich_and_clean_aba3, enrich_and_clean_aba4,
    to_json_records
)

from app.utils.costs.variable.meli.config import (
    SHEET_CANDS_TARIFAS_FULL_ARMAZEN,
    SHEET_CANDS_TARIFAS_FULL_RETIRADA,
    SHEET_CANDS_TARIFAS_FULL_COLETA,
    SHEET_CANDS_TARIFAS_FULL_ARMAZEN_PROL,
    resolve_sheet_name,
)

def _available_sheets(xlsx_path: str | bytes) -> List[str]:
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception:
        return []

def _read_resolved(xlsx_path: str | bytes, sheet_candidates: List[str], header_row: int) -> pd.DataFrame:
    """
    Resolve o nome da aba por candidatos (centralizados em meli/config.py).
    Se não encontrar, retorna DF vazio (mês sem esse custo → soma 0).
    """
    sheets = _available_sheets(xlsx_path)
    name = resolve_sheet_name(sheets, sheet_candidates)
    if not name:
        warnings.warn(
            f"[tarifas_full] Aba não encontrada neste mês; tratando como vazio. "
            f"Procurado≈{sheet_candidates} | Encontradas={sheets}"
        )
        return pd.DataFrame()
    return pd.read_excel(
        xlsx_path, sheet_name=name, header=header_row,
        dtype=object, engine="openpyxl"
    )

def read_tarifa_armazenamento(xlsx_path: str | bytes, competencia: Optional[str] = None, header_row: int = 5):
    df = _read_resolved(xlsx_path, SHEET_CANDS_TARIFAS_FULL_ARMAZEN, header_row)

    df = map_aba1(df)
    df = enrich_and_clean_aba1(df, competencia)
    return to_json_records(df)

def read_custo_retirada_estoque(xlsx_path: str | bytes, competencia: Optional[str] = None, header_row: int = 5):
    df = _read_resolved(xlsx_path, SHEET_CANDS_TARIFAS_FULL_RETIRADA, header_row)

    df = map_aba2(df)
    df = enrich_and_clean_aba2(df, competencia)
    return to_json_records(df)

def read_custo_servico_coleta(xlsx_path: str | bytes, competencia: Optional[str] = None, header_row: int = 5):
    df = _read_resolved(xlsx_path, SHEET_CANDS_TARIFAS_FULL_COLETA, header_row)
    df = map_aba3(df)
    df = enrich_and_clean_aba3(df, competencia)
    return to_json_records(df)

def read_custo_armazenamento_prolongado(xlsx_path: str | bytes, competencia: Optional[str] = None, header_row: int = 5):
    df = _read_resolved(xlsx_path, SHEET_CANDS_TARIFAS_FULL_ARMAZEN_PROL, header_row)
    df = map_aba4(df)
    df = enrich_and_clean_aba4(df, competencia)
    return to_json_records(df)
